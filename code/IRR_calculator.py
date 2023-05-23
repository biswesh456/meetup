import openpyxl

if __name__ == "__main__":
    total_correct = 0
    total_length = 0

    for idx in range(4,12):
        wb1 = openpyxl.load_workbook("seemab_dial_"+str(idx)+".xlsx")
        ws1 = wb1.active

        seemab_annotations = []

        for row in ws1.iter_rows(values_only=True):
            if row[2] != None:
                if 'A:' in row[2] or 'B:' in row[2]:
                    # print(row[4].lower())
                    seemab_annotations.append(row[4].lower().strip())

        wb2 = openpyxl.load_workbook("biswesh_dial_"+str(idx)+".xlsx")
        ws2 = wb2.active

        biswesh_annotations = []

        print("----")

        for row in ws2.iter_rows(values_only=True):
            if row[2] != None:
                if 'A:' in row[2] or 'B:' in row[2]:
                    # print(row[4].lower()) 
                    biswesh_annotations.append(row[4].lower().strip())


        zipped_annotations = zip(biswesh_annotations, seemab_annotations)

        # print('zipped')
        
        matching_count = 0
        for i,z in enumerate(zipped_annotations):
            if z[0] == z[1]:
                matching_count += 1
            else:
                print(i, z)

        total_correct += matching_count
        total_length += len(biswesh_annotations)
        irr = matching_count/len(biswesh_annotations)
        print("IRR for dialog "+ str(idx)+ ": ", irr)

    final_irr = total_correct/total_length
    print("Finall IRR : ", final_irr)